# Finance tracker

import sqlite3
import os
import datetime
import csv
import openpyxl
from fpdf import FPDF

script_directory = os.path.dirname(os.path.abspath(__file__))
db_path = os.path.join(script_directory, 'finance.db')

exports_path = os.path.join(script_directory, 'exports')
if not os.path.exists(exports_path):
    os.makedirs(exports_path)

def init_db():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY,
            date TEXT,
            category TEXT,
            description TEXT,
            amount REAL

        )
    ''')
    connect_to_database.commit()
    connect_to_database.close()

def insert_transaction():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    
    while True:
        date = input("Enter the date (DD-MM-YYYY): ")
        try:
            datetime.datetime.strptime(date, "%d-%m-%Y")
            break
        except ValueError:
            print("Invalid input. Please use the format DD-MM-YYYY (e.g. 31-01-2025).")

    category = input("Enter the category (e.g. Food, Bills): ")
    description = input("Description: ")
    amount = float(input("Enter the amount. In case of decimals use dot(.) instead of comma(,): "))

    db_cursor.execute('''
        INSERT INTO transactions (date, category, description, amount)
        VALUES (?, ?, ?, ?)
    ''', (date, category, description, amount))

    connect_to_database.commit()
    connect_to_database.close()

def view_all_transactions():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    print("\n=== Transaction history ===")
    for row in rows:
        print(row)
    connect_to_database.close()

def view_transactions_by_month():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    month_year = input("Enter month and year (MM-YYYY): ")
    search_pattern = f'%-{month_year}'
    db_cursor.execute('SELECT * FROM transactions WHERE date LIKE ?', (search_pattern,))
    rows = db_cursor.fetchall()
    print(f"\n=== Transaction history for {month_year} ===")
    if not rows:
        print("No transactions found for this period.")
    else:
        for row in rows:
            print(row)
    connect_to_database.close()

def view_transactions_by_week():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    week_year = input("Enter the week and the year (WW-YYYY):  ")
    try:
        week, year = map(int, week_year.split('-'))
        start_date = datetime.datetime.strptime(f'{year}-W{week - 1}-1', "%Y-W%W-%w").date()
        end_date = start_date + datetime.timedelta(days=6)
    except Exception:
        print("Invalid input. Please use the format WW-YYYY (e.g. 12-2025).")
        return
    
    start_str = start_date.strftime("%d-%m-%Y")
    end_str = end_date.strftime("%d-%m-%Y")

    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    filtered = []
    for row in rows:
        try:
            row_date = datetime.datetime.strptime(row[1], "%d-%m-%Y").date()
            if start_date <= row_date <= end_date:
                filtered.append(row)
        except Exception:
            continue

    print(f"\n=== Transaction history for week {week} of {year} ({start_str} to {end_str}) ===")
    if not filtered:
        print("No transactions found for this week.")
    else:
        for row in filtered:
            print(row)
    connect_to_database.close()

def clear_all_transactions():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('DELETE FROM transactions')
    connect_to_database.commit()
    connect_to_database.execute('VACUUM')
    connect_to_database.close()
    print("All transactions have been deleted and the database file has been compacted.")

def export_transactions_to_csv():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    connect_to_database.close()

    filename = input("Enter filename for CSV export (e.g. transactions.csv): ")
    if not filename.endswith('.csv'):
        filename += '.csv'
    filepath = os.path.join(exports_path, filename)

    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, delimiter=';') # ';' is used as the delimiter to make differentiating categories easier.
        writer.writerow(['ID', 'Date', 'Category', 'Description', 'Amount'])
        writer.writerows(rows)
    print(f"Transactions have been exported to {filepath}")

def export_transactions_to_excel():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    connect_to_database.close()

    filename = input("Enter filename for Excel export (e.g. transactions.xlsx): ")
    if not filename.endswith('xlsx'):
        filename += '.xlsx'
    filepath = os.path.join(exports_path, filename)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(['ID', 'Date', 'Category', 'Description', 'Amount'])
    for row in rows:
        worksheet.append(row)
    workbook.save(filepath)
    print(f"Transactions have been exported to {filepath}")

def export_transactions_to_pdf():
    connect_to_database = sqlite3.connect(db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    connect_to_database.close()

    filename = input("Enter filename for PDF export (e.g, transactions.pdf): ")
    if not filename.endswith('.pdf'):
        filename += '.pdf'
    filepath = os.path.join(exports_path, filename)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Transaction History", ln=True, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=10)
    headers = ['ID', 'Date', 'Category', 'Description', 'Amount']
    pdf.cell(10, 10, headers[0], 1)
    pdf.cell(30, 10, headers[1], 1)
    pdf.cell(30, 10, headers[2], 1)
    pdf.cell(80, 10, headers[3], 1)
    pdf.cell(30, 10, headers[4], 1)
    pdf.ln()
    for row in rows:
        pdf.cell(10, 10, str(row[0]), 1)
        pdf.cell(30, 10, row[1], 1)
        pdf.cell(30, 10, row[2], 1)
        pdf.cell(80, 10, row[3], 1)
        pdf.cell(30, 10, str(row[4]), 1)
        pdf.ln()
    pdf.output(filepath)
    print(f"Transactions have been exported to {filepath}")

init_db()

while True:
    print("\n1. Add transaction")
    print("2. View all transactions")
    print("3. View transactions by month")
    print("4. View transactions by week")
    print("5. Export transactions")
    print("6. Clear all transactions")
    print("7. Exit")
    choice = input("Choose option: ")

    if choice == '1':
        insert_transaction()
    elif choice == '2':
        view_all_transactions()
    elif choice == '3':
        view_transactions_by_month()
    elif choice == '4':
        view_transactions_by_week()
    elif choice == '5':
        print("Export data as 1) CSV, 2) Excel, 3) PDF")
        exportchoice = input("Choose export format: ")
        if exportchoice == '1':
            export_transactions_to_csv()
        elif exportchoice == '2':
            export_transactions_to_excel()
        elif exportchoice == '3':
            export_transactions_to_pdf()
    elif choice == '6':
        print("Are you sure you want to delete all of your data?")
        final_choice = input("If you are, write 'delete my data'. If not, input the letter 'n' and press enter: ")
        if final_choice.lower() == 'delete my data':
            clear_all_transactions()
        else:
            print("Your data has not been removed.")
            continue
    elif choice == '7':
        break
    else:
        print("Invalid input.")