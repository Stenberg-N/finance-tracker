import sqlite3
from pathlib import Path
import csv
import openpyxl
from fpdf import FPDF
from app.config import EXPORTS_PATH, DB_PATH
from cryptography.fernet import Fernet

def export_transactions_to_csv(user_id, filename):
    from database.db import encryption_key
    fernet = Fernet(encryption_key)
    connect_to_database = sqlite3.connect(DB_PATH)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT id, date, category, description, amount, type FROM transactions WHERE user_id = ?', (user_id,))
    rows = db_cursor.fetchall()
    connect_to_database.close()

    decrypted_rows = []
    for row in rows:
        decrypted_date = fernet.decrypt(row[1].encode()).decode()
        decrypted_category = fernet.decrypt(row[2].encode()).decode()
        decrypted_description = fernet.decrypt(row[3].encode()).decode()
        decrypted_amount = float(fernet.decrypt(row[4].encode()).decode())
        decrypted_rows.append((row[0], decrypted_date, decrypted_category, decrypted_description, decrypted_amount, row[5]))

    if not filename.endswith('.csv'):
        filename += '.csv'
    filepath = EXPORTS_PATH / filename
    EXPORTS_PATH.mkdir(exist_ok=True)

    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, delimiter=';')
        writer.writerow(['ID', 'Date', 'Category', 'Description', 'Amount', 'Type'])
        for row in decrypted_rows:
            writer.writerow(row)

    return decrypted_rows

def export_transactions_to_excel(user_id, filename):
    from database.db import encryption_key
    fernet = Fernet(encryption_key)
    connect_to_database = sqlite3.connect(DB_PATH)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT id, date, category, description, amount, type FROM transactions WHERE user_id = ?', (user_id,))
    rows = db_cursor.fetchall()
    connect_to_database.close()

    decrypted_rows = []
    for row in rows:
        decrypted_date = fernet.decrypt(row[1].encode()).decode()
        decrypted_category = fernet.decrypt(row[2].encode()).decode()
        decrypted_description = fernet.decrypt(row[3].encode()).decode()
        decrypted_amount = float(fernet.decrypt(row[4].encode()).decode())
        decrypted_rows.append((row[0], decrypted_date, decrypted_category, decrypted_description, decrypted_amount, row[5]))

    if not filename.endswith('xlsx'):
        filename += '.xlsx'
    filepath = EXPORTS_PATH / filename
    EXPORTS_PATH.mkdir(exist_ok=True)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(['ID', 'Date', 'Category', 'Description', 'Amount', 'Type'])
    for row in decrypted_rows:
        worksheet.append(row)
    workbook.save(filepath)

    return decrypted_rows

def export_transactions_to_pdf(user_id, filename):
    from database.db import encryption_key
    fernet = Fernet(encryption_key)
    connect_to_database = sqlite3.connect(DB_PATH)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT id, date, category, description, amount, type FROM transactions WHERE user_id = ?', (user_id,))
    rows = db_cursor.fetchall()
    connect_to_database.close()

    decrypted_rows = []
    for row in rows:
        decrypted_date = fernet.decrypt(row[1].encode()).decode()
        decrypted_category = fernet.decrypt(row[2].encode()).decode()
        decrypted_description = fernet.decrypt(row[3].encode()).decode()
        decrypted_amount = float(fernet.decrypt(row[4].encode()).decode())
        decrypted_rows.append((row[0], decrypted_date, decrypted_category, decrypted_description, decrypted_amount, row[5]))

    if not filename.endswith('.pdf'):
        filename += '.pdf'
    filepath = EXPORTS_PATH / filename
    EXPORTS_PATH.mkdir(exist_ok=True)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Transaction History", ln=True, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=10)
    headers = ['ID', 'Date', 'Category', 'Description', 'Amount', 'Type']
    pdf.cell(10, 10, headers[0], 1)
    pdf.cell(30, 10, headers[1], 1)
    pdf.cell(30, 10, headers[2], 1)
    pdf.cell(50, 10, headers[3], 1)
    pdf.cell(30, 10, headers[4], 1)
    pdf.cell(30, 10, headers[5], 1)
    pdf.ln()
    for row in decrypted_rows:
        pdf.cell(10, 10, str(row[0]), 1)
        pdf.cell(30, 10, row[1], 1)
        pdf.cell(30, 10, row[2], 1)
        pdf.cell(50, 10, row[3], 1)
        pdf.cell(30, 10, str(row[4]), 1)
        pdf.cell(30, 10, row[5], 1)
        pdf.ln()
    pdf.output(filepath)

    return decrypted_rows