# Export logic

import sqlite3
import os
import csv
import openpyxl
from fpdf import FPDF
import config

def export_transactions_to_csv(filename):
    connect_to_database = sqlite3.connect(config.db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    connect_to_database.close()

    if not filename.endswith('.csv'):
        filename += '.csv'
    filepath = os.path.join(config.exports_path, filename)

    with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, delimiter=';') # ';' is used as the delimiter to make differentiating categories easier.
        writer.writerow(['ID', 'Date', 'Category', 'Description', 'Amount', 'Type'])
        writer.writerows(rows)

def export_transactions_to_excel(filename):
    connect_to_database = sqlite3.connect(config.db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    connect_to_database.close()

    if not filename.endswith('xlsx'):
        filename += '.xlsx'
    filepath = os.path.join(config.exports_path, filename)

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(['ID', 'Date', 'Category', 'Description', 'Amount', 'Type'])
    for row in rows:
        worksheet.append(row)
    workbook.save(filepath)

def export_transactions_to_pdf(filename):
    connect_to_database = sqlite3.connect(config.db_path)
    db_cursor = connect_to_database.cursor()
    db_cursor.execute('SELECT * FROM transactions')
    rows = db_cursor.fetchall()
    connect_to_database.close()

    if not filename.endswith('.pdf'):
        filename += '.pdf'
    filepath = os.path.join(config.exports_path, filename)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Transaction History", ln=True, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=10)
    headers = ['ID', 'Date', 'Category', 'Description', 'Amount', 'Type']
    pdf.cell(10, 10, headers[0], 1)
    pdf.cell(30, 10, headers[1], 1)
    pdf.cell(30, 10, headers[2], 1)
    pdf.cell(50, 10, headers[3], 1)
    pdf.cell(30, 10, headers[4], 1)
    pdf.cell(30, 10, headers[5], 1)
    pdf.ln()
    for row in rows:
        pdf.cell(10, 10, str(row[0]), 1)
        pdf.cell(30, 10, row[1], 1)
        pdf.cell(30, 10, row[2], 1)
        pdf.cell(50, 10, row[3], 1)
        pdf.cell(30, 10, str(row[4]), 1)
        pdf.cell(30, 10, row[5], 1)
        pdf.ln()
    pdf.output(filepath)